from openpyxl import load_workbook

wb = load_workbook("sessiontrack2014_full.xlsx")
file = wb.worksheets
f = open('data.txt', 'w')
sessionData = file[0]
maxRow = sessionData.max_row
print("max: ", maxRow)
f.write("<topics>")
f.write("\n")
f.write("\n")


for i in range(2, maxRow, +1):
    sessionID = str(sessionData.cell(i, 2).value)
    queryID = str(sessionData.cell(i, 9).value)
    title = str(sessionData.cell(i, 18).value)
    if(title == str(sessionData.cell(i+1, 18).value)):
        i = i+1
    else:
        f.write("<top>" + "\n")
        f.write("\n")
        f.write("<num>" + "s" + sessionID + "-" + "q" + queryID + "</num>" + "\n")
        f.write("<title>" + title + "</title>" + "\n")
        f.write("\n")
        f.write("<desc>" + title + "</desc>" + "\n")
        f.write("\n")
        f.write("<narr>" + title + "</narr>" + "\n")
        f.write("</top>" + "\n")
        f.write("\n")

f.close()






# # 获取第一行所有数据
# row1 = []
# print(sessionData[1])
# for row in sessionData[1]:
#     print(row)
#     row1.append(row.value)
# print(row1)
#
# # 获取第一列所有数据
# col1 = []
# for col in sessionData['A']:
#     col1.append(col.value)
# print(col1)
