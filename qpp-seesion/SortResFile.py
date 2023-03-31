from openpyxl import load_workbook

wb = load_workbook("qrels.xlsx")
file = wb.worksheets
f = open('data2.txt', 'w')
qrelsData = file[0]
maxRow = qrelsData.max_row

for i in range(2, maxRow, +1):
    sessionID = str(qrelsData.cell(i, 1).value)
    queryID = str(qrelsData.cell(i, 2).value)
    docID = str(qrelsData.cell(i, 3).value)
    sim = str(qrelsData.cell(i, 4).value)
    rank = str(qrelsData.cell(i, 5).value)
    name = "BM25"
    f.write(sessionID + " " + "Q" + queryID + " " + docID + " " + rank + " " + sim + "\n")

f.close()

