from openpyxl import load_workbook
from collections import defaultdict
dict = defaultdict(list)
wb = load_workbook("SessionQueryScore.xlsx")
file = wb.worksheets
f = open('data_avg.txt', 'w')
qppData = file[0]
maxRow = qppData.max_row
total_score = 0
queryScoreList = []
dict1 = defaultdict(list)
for i in range(1, maxRow+1, +1):
    sessionQueryID = str(qppData.cell(i, 1).value)
    queryScore = qppData.cell(i, 2).value
    s = sessionQueryID.split('-')
    id = s[0]

    if id in sessionQueryID:
        total_score += total_score + queryScore
        dict[id].append(queryScore)

dict_avg = {}
for key in dict.keys():
    avg = sum(dict[key])/len(dict[key])
    dict_avg[key] = avg
    f.write(str(dict_avg[key]) + "\n")

print(dict_avg)


f.close()

